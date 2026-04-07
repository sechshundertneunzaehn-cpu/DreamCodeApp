#!/usr/bin/env python3
"""
Erstellt DreamCodeApp_Forschungsdaten_Komplett.xlsx
Quelle: sddb_full_analysis.json, sddb_studies_research.json, geo_complete.json, sddb_dreams.csv
"""

import json
import csv
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl nicht gefunden. Installation mit: pip install openpyxl --break-system-packages")
    sys.exit(1)

# ── Pfade ──────────────────────────────────────────────────────────────────────
BASE = Path("/home/dejavu/DreamCodeApp/data")
FULL_ANALYSIS  = BASE / "analysis" / "sddb_full_analysis.json"
STUDIES        = BASE / "analysis" / "sddb_studies_research.json"
GEO            = BASE / "analysis" / "geo_complete.json"
CSV_FILE       = BASE / "raw" / "sddb_dreams.csv"
OUTPUT         = BASE / "analysis" / "DreamCodeApp_Forschungsdaten_Komplett.xlsx"

# ── Styling ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F3864")   # dunkelblau
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
NUM_ALIGN    = Alignment(horizontal="right")
DEF_ALIGN    = Alignment(horizontal="left", wrap_text=False)

def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN

def auto_width(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def freeze_header(ws):
    ws.freeze_panes = "A2"

# ── Daten laden ────────────────────────────────────────────────────────────────
print("Lade JSON-Dateien …")
with open(FULL_ANALYSIS) as f:
    full = json.load(f)
with open(STUDIES) as f:
    studies_data = json.load(f)
with open(GEO) as f:
    geo = json.load(f)

# ── Workbook ───────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Default-Sheet entfernen

# ══════════════════════════════════════════════════════════════════════════════
# Blatt 1: Studien_Übersicht
# ══════════════════════════════════════════════════════════════════════════════
print("Blatt 1: Studien_Übersicht …")
ws1 = wb.create_sheet("Studien_Übersicht")

headers1 = ["ID", "Studie_Name", "Jahr", "Hauptforscher", "Mitwirkende",
            "Institution", "Land", "Anzahl_Berichte", "DOI_Publikation", "URL", "Beschreibung"]
ws1.append(headers1)

row_id = 1

# Aus sddb_studies_research.json → studies[]
for study in studies_data.get("studies", []):
    researchers = study.get("researchers", [])
    lead = researchers[0] if researchers else ""
    collab = ", ".join(researchers[1:]) if len(researchers) > 1 else ""

    doi = study.get("doi") or ""
    url = (study.get("doi_url") or study.get("publication_url") or
           study.get("pdf_url") or study.get("pubmed") or "")

    year = study.get("year") or study.get("published_year") or ""

    ws1.append([
        f"S-{row_id:03d}",
        study.get("name", ""),
        str(year),
        lead,
        collab,
        study.get("institution", ""),
        study.get("country", ""),
        study.get("sample_size") or "",
        doi,
        url,
        study.get("description", ""),
    ])
    row_id += 1

# Dryad-Studie aus studies_data
dryad = studies_data.get("dryad", {})
if dryad:
    researchers = dryad.get("researchers", [])
    lead = researchers[0].get("name", "") if researchers else ""
    collab = ", ".join(r.get("name", "") for r in researchers[1:])
    affils = "; ".join(r.get("affiliation", "") for r in researchers)
    ws1.append([
        f"S-{row_id:03d}",
        dryad.get("name", ""),
        str(dryad.get("year", "")),
        lead,
        collab,
        affils,
        "International",
        "",
        dryad.get("doi", ""),
        dryad.get("doi_url", ""),
        dryad.get("description", "") if dryad.get("description") else
            "Open-access dream reports dataset published on Dryad.",
    ])
    row_id += 1

# Aus sddb_full_analysis.json → surveys{} – je Survey eine Zeile
db = studies_data.get("database", {})
for survey_name, sdata in full.get("surveys", {}).items():
    ws1.append([
        f"S-{row_id:03d}",
        survey_name,
        "",   # Jahr nicht in full_analysis
        db.get("director", "Kelly Bulkeley"),
        "",
        db.get("institution", "Sleep and Dream Database"),
        "USA",
        sdata.get("count", ""),
        "",
        db.get("url", "https://sleepanddreamdatabase.org/"),
        f"SDDb Survey mit {sdata.get('count', 0)} Einträgen.",
    ])
    row_id += 1

style_header_row(ws1)
freeze_header(ws1)
auto_width(ws1)

# ══════════════════════════════════════════════════════════════════════════════
# Blatt 2: Geografische_Verteilung
# ══════════════════════════════════════════════════════════════════════════════
print("Blatt 2: Geografische_Verteilung …")
ws2 = wb.create_sheet("Geografische_Verteilung")

headers2 = ["ID", "Land", "ISO_Code", "Anzahl_Berichte", "Anteil_Prozent",
            "US_States_vorhanden", "Top_States", "Koordinaten_Lat", "Koordinaten_Lng"]
ws2.append(headers2)

# ISO-Mapping für bekannte Länder
ISO_MAP = {
    "United States": "US", "Russia": "RU", "Brazil": "BR", "Japan": "JP",
    "England": "GB", "Argentina": "AR", "Ukraine": "UA", "Australia": "AU",
    "Canada": "CA", "New Zealand": "NZ", "Sweden": "SE", "India": "IN",
    "Ireland": "IE", "Philippines": "PH", "Denmark": "DK", "South Africa": "ZA",
    "Israel": "IL", "Poland": "PL", "Bangladesh": "BD", "Scotland": "GB",
    "Northern Ireland": "GB", "Greece": "GR", "Cyprus": "CY", "Mexico": "MX",
    "Albania": "AL", "Malaysia": "MY", "Belgium": "BE", "Thailand": "TH",
    "Netherlands": "NL", "Saudi Arabia": "SA", "Pakistan": "PK", "Malta": "MT",
    "Zimbabwe": "ZW", "Nepal": "NP",
}

# Koordinaten der Hauptstädte/Mittelpunkte
COUNTRY_COORDS = {
    "United States": (37.09, -95.71), "Russia": (61.52, 105.32),
    "Brazil": (-14.24, -51.93), "Japan": (36.2, 138.25),
    "England": (52.36, -1.17), "Argentina": (-38.42, -63.62),
    "Ukraine": (48.38, 31.17), "Australia": (-25.27, 133.78),
    "Canada": (56.13, -106.35), "New Zealand": (-40.9, 174.89),
    "Sweden": (60.13, 18.64), "India": (20.59, 78.96),
    "Ireland": (53.41, -8.24), "Philippines": (12.88, 121.77),
    "Denmark": (56.26, 9.5), "South Africa": (-30.56, 22.94),
    "Israel": (31.05, 34.85), "Poland": (51.92, 19.15),
    "Bangladesh": (23.69, 90.35), "Scotland": (56.49, -4.2),
    "Northern Ireland": (54.61, -6.69), "Greece": (39.07, 21.82),
    "Cyprus": (35.13, 33.43), "Mexico": (23.63, -102.55),
    "Albania": (41.15, 20.17), "Malaysia": (4.21, 101.98),
    "Belgium": (50.5, 4.47), "Thailand": (15.87, 100.99),
    "Netherlands": (52.13, 5.29), "Saudi Arabia": (23.89, 45.08),
    "Pakistan": (30.38, 69.35), "Malta": (35.94, 14.37),
    "Zimbabwe": (-19.02, 29.15), "Nepal": (28.39, 84.12),
}

by_country = geo["sddb_geographic"]["by_country"]
total_reports = sum(v["count"] for v in by_country.values())

for idx, (country, cdata) in enumerate(by_country.items(), start=1):
    count = cdata["count"]
    pct = round(count / total_reports * 100, 2) if total_reports else 0
    states = cdata.get("states", {})
    has_states = "Ja" if states else "Nein"
    top_states = ", ".join(list(states.keys())[:5]) if states else ""
    iso = ISO_MAP.get(country, "")
    lat, lng = COUNTRY_COORDS.get(country, ("", ""))

    ws2.append([
        f"G-{idx:03d}",
        country,
        iso,
        count,
        pct,
        has_states,
        top_states,
        lat,
        lng,
    ])

style_header_row(ws2)
freeze_header(ws2)
auto_width(ws2)

# ══════════════════════════════════════════════════════════════════════════════
# Blatt 3: US_States
# ══════════════════════════════════════════════════════════════════════════════
print("Blatt 3: US_States …")
ws3 = wb.create_sheet("US_States")

headers3 = ["ID", "State", "Anzahl_Berichte", "Anteil_Prozent", "Top_ZIP_Codes", "Region"]
ws3.append(headers3)

STATE_REGION = {
    "California": "West", "Oregon": "West", "Washington": "West", "Nevada": "West",
    "Arizona": "West", "Utah": "West", "Colorado": "West", "Idaho": "West",
    "Montana": "West", "Wyoming": "West", "New Mexico": "West", "Hawaii": "West",
    "Alaska": "West",
    "Texas": "South", "Florida": "South", "Georgia": "South", "North Carolina": "South",
    "Virginia": "South", "South Carolina": "South", "Tennessee": "South",
    "Alabama": "South", "Louisiana": "South", "Oklahoma": "South",
    "Arkansas": "South", "Mississippi": "South", "Kentucky": "South",
    "West Virginia": "South", "Maryland": "South", "Delaware": "South",
    "Washington DC": "South",
    "New York": "Northeast", "Pennsylvania": "Northeast", "New Jersey": "Northeast",
    "Massachusetts": "Northeast", "Connecticut": "Northeast", "New Hampshire": "Northeast",
    "Vermont": "Northeast", "Rhode Island": "Northeast", "Maine": "Northeast",
    "Illinois": "Midwest", "Ohio": "Midwest", "Michigan": "Midwest",
    "Indiana": "Midwest", "Wisconsin": "Midwest", "Minnesota": "Midwest",
    "Iowa": "Midwest", "Missouri": "Midwest", "Kansas": "Midwest",
    "Nebraska": "Midwest", "South Dakota": "Midwest", "North Dakota": "Midwest",
}

by_state = geo["sddb_geographic"]["by_state"]
us_total = sum(v["count"] for v in by_state.values())

for idx, (state, sdata) in enumerate(sorted(by_state.items(), key=lambda x: -x[1]["count"]), start=1):
    count = sdata["count"]
    pct = round(count / us_total * 100, 2) if us_total else 0
    zips = ", ".join(sdata.get("zip_codes", [])[:5])
    region = STATE_REGION.get(state, "")
    ws3.append([
        f"US-{idx:03d}",
        state,
        count,
        pct,
        zips,
        region,
    ])

style_header_row(ws3)
freeze_header(ws3)
auto_width(ws3)

# ══════════════════════════════════════════════════════════════════════════════
# Blatt 4: Demografische_Daten
# ══════════════════════════════════════════════════════════════════════════════
print("Blatt 4: Demografische_Daten …")
ws4 = wb.create_sheet("Demografische_Daten")

headers4 = ["ID", "Kategorie", "Wert", "Anzahl", "Anteil_Prozent"]
ws4.append(headers4)

demographics = full.get("demographics", {})
demo_row_id = 1

CATEGORY_MAP = {
    "Race/Ethnicity": "Race_Ethnicity",
    "Race/Ethnicity D": "Race_Ethnicity_D",
    "Race/Ethnicity E": "Race_Ethnicity_E",
    "Age Range": "Age_Range",
    "Age": "Age",
    "Sex Assigned at Birth": "Sex",
    "Religious Affiliation A": "Religion_A",
    "Religious Affiliation B": "Religion_B",
    "Education Level": "Education",
    "Income Tier": "Income",
    "Income (values)": "Income_Values",
    "Political Party": "Political_Party",
    "Political Ideology": "Political_Ideology",
    "Marital status": "Marital_Status",
    "Marital Status A": "Marital_Status_A",
    "Generation": "Generation",
    "Dream Recall": "Dream_Recall",
    "Nightmare Recall": "Nightmare_Recall",
    "Hours of sleep per night": "Sleep_Hours",
}

for field, values_dict in demographics.items():
    cat_label = CATEGORY_MAP.get(field, field)
    vals = values_dict.get("values", {})
    total_cat = sum(vals.values()) if vals else 0
    for val, count in sorted(vals.items(), key=lambda x: -x[1]):
        pct = round(count / total_cat * 100, 2) if total_cat else 0
        ws4.append([
            f"D-{demo_row_id:04d}",
            cat_label,
            str(val),
            count,
            pct,
        ])
        demo_row_id += 1

style_header_row(ws4)
freeze_header(ws4)
auto_width(ws4)

# ══════════════════════════════════════════════════════════════════════════════
# Blatt 5: Traumberichte_Komplett
# ══════════════════════════════════════════════════════════════════════════════
print("Blatt 5: Traumberichte_Komplett (39.089 Einträge – das dauert …)")
ws5 = wb.create_sheet("Traumberichte_Komplett")

headers5 = ["ID", "Text_Vorschau_100Zeichen", "Wortanzahl", "Survey", "Respondent_ID",
            "Land", "US_State", "Region", "ZIP", "Alter", "Geschlecht",
            "Ethnizität", "Dream_Location", "Datum"]
ws5.append(headers5)

# Spaltenindizes aus CSV-Header ermitteln
COLS_OF_INTEREST = {
    "answer_text":              None,
    "word_count":               None,
    "survey":                   None,
    "respondent":               None,
    "Country of Residence":     None,
    "US State of Residence":    None,
    "Region":                   None,
    "ZIP Code":                 None,
    "Age Range":                None,
    "Sex Assigned at Birth":    None,
    "Race/Ethnicity":           None,
    "Dream Location":           None,
    "date":                     None,
    "In what country or region do you currently reside?": None,
}

dream_count = 0
with open(CSV_FILE, encoding="utf-8") as f:
    reader = csv.reader(f)
    header_row = next(reader)

    # Indices zuweisen
    for key in COLS_OF_INTEREST:
        try:
            COLS_OF_INTEREST[key] = header_row.index(key)
        except ValueError:
            pass  # Spalte existiert nicht

    def get(row, col):
        if COLS_OF_INTEREST.get(col) is not None:
            idx = COLS_OF_INTEREST[col]
            if idx < len(row):
                return row[idx].strip()
        return ""

    for row in reader:
        dream_count += 1
        text = get(row, "answer_text")
        preview = text[:100] if text else ""
        wc = get(row, "word_count")
        survey = get(row, "survey")
        resp = get(row, "respondent")
        country = get(row, "Country of Residence") or get(row, "In what country or region do you currently reside?")
        state = get(row, "US State of Residence")
        region = get(row, "Region")
        zip_code = get(row, "ZIP Code")
        age = get(row, "Age Range")
        sex = get(row, "Sex Assigned at Birth")
        ethnicity = get(row, "Race/Ethnicity")
        dream_loc = get(row, "Dream Location")
        date = get(row, "date")

        ws5.append([
            f"DR-{dream_count:05d}",
            preview,
            int(wc) if wc and wc.isdigit() else (wc or ""),
            survey,
            resp,
            country,
            state,
            region,
            zip_code,
            age,
            sex,
            ethnicity,
            dream_loc,
            date,
        ])

        if dream_count % 5000 == 0:
            print(f"  … {dream_count} Einträge verarbeitet")

print(f"  Gesamt: {dream_count} Einträge")

style_header_row(ws5)
freeze_header(ws5)
# Für 39k+ Zeilen nur feste Breite setzen (auto_width wäre zu langsam)
FIXED_WIDTHS = [10, 55, 12, 35, 18, 20, 20, 15, 10, 12, 12, 25, 20, 22]
for i, w in enumerate(FIXED_WIDTHS, start=1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# Blatt 6: Bot_Profile_Geo
# ══════════════════════════════════════════════════════════════════════════════
print("Blatt 6: Bot_Profile_Geo …")
ws6 = wb.create_sheet("Bot_Profile_Geo")

headers6 = ["ID", "Land_ISO", "Land_Name", "Stadt", "Lat", "Lng",
            "Bots_in_Land", "Staedte_Anzahl"]
ws6.append(headers6)

# ISO → Ländernames
ISO_TO_NAME = {v: k for k, v in ISO_MAP.items()}
ISO_TO_NAME.update({
    "DE": "Deutschland", "IN": "Indien", "BR": "Brasilien", "GB": "Großbritannien",
    "FR": "Frankreich", "MX": "Mexiko", "JP": "Japan", "ES": "Spanien",
    "IT": "Italien", "AU": "Australien", "KR": "Südkorea", "TR": "Türkei",
    "KZ": "Kasachstan", "NL": "Niederlande", "CA": "Kanada", "ID": "Indonesien",
    "CN": "China", "SE": "Schweden", "PL": "Polen", "CO": "Kolumbien",
    "PH": "Philippinen", "TH": "Thailand", "UZ": "Usbekistan", "AR": "Argentinien",
    "VN": "Vietnam", "EG": "Ägypten", "NG": "Nigeria", "NZ": "Neuseeland",
    "ZA": "Südafrika", "MA": "Marokko", "AE": "Vereinigte Arabische Emirate",
    "KE": "Kenia",
})

bot_by_country = geo["bot_profiles"]["by_country"]
cities_by_country = geo["generate_bots_cities"]["by_country"]
bot_row_id = 1

for iso, bdata in bot_by_country.items():
    count = bdata["count"]
    cities_list = bdata.get("cities", cities_by_country.get(iso, []))
    country_name = ISO_TO_NAME.get(iso, iso)
    num_cities = len(cities_list)

    for city in cities_list:
        ws6.append([
            f"B-{bot_row_id:04d}",
            iso,
            country_name,
            city.get("name", ""),
            city.get("lat", ""),
            city.get("lng", ""),
            count,
            num_cities,
        ])
        bot_row_id += 1

style_header_row(ws6)
freeze_header(ws6)
auto_width(ws6)

# ══════════════════════════════════════════════════════════════════════════════
# Blatt 7: Datenquellen_Referenzen
# ══════════════════════════════════════════════════════════════════════════════
print("Blatt 7: Datenquellen_Referenzen …")
ws7 = wb.create_sheet("Datenquellen_Referenzen")

headers7 = ["ID", "Quelle", "Typ", "URL", "DOI", "Forscher", "Institution",
            "Beschreibung", "Status"]
ws7.append(headers7)

refs = [
    ["REF-001", "Sleep and Dream Database (SDDb)", "Datenbank",
     "https://sleepanddreamdatabase.org/", "",
     "Kelly Bulkeley, Ph.D.", "Graduate Theological Union, Berkeley, CA",
     "Open-access digitales Archiv mit 44.500+ Traumberichten. Enthält 18 demografische Surveys, Suchfunktionen und Wortanalyse-Tools.",
     "Aktiv"],
    ["REF-002", "SDDb Zenodo Mirror", "Datensatz (Zenodo)",
     "https://zenodo.org/records/11662064", "10.5281/zenodo.11662064",
     "Kelly Bulkeley", "SDDb / Zenodo",
     "Vollständiger Export der SDDb-Daten als CSV auf Zenodo (öffentlich zugänglich).",
     "Aktiv"],
    ["REF-003", "Dryad: Our Dreams, Our Selves", "Datensatz (Dryad)",
     "https://datadryad.org/dataset/doi:10.5061/dryad.qbzkh18fr",
     "10.5061/dryad.qbzkh18fr",
     "Luca Maria Aiello, Daniele Quercia, Alessandro Fogli",
     "Nokia Bell Labs / King's College London",
     "Automatische Interpretation von Traumberichten; enthält dream-reports aus DreamBank.net.",
     "Archiviert 2020"],
    ["REF-004", "DreamBank.net", "Datenbank",
     "https://dreambank.net/", "",
     "G. William Domhoff, Adam Schneider", "UC Santa Cruz",
     "Online-Suchmaschine für Traumberichte; Heimat der Hall-Van de Castle Normen und Barb Sanders Serie.",
     "Aktiv"],
    ["REF-005", "DreamResearch.net", "Forschungswebsite",
     "https://dreams.ucsc.edu/", "",
     "G. William Domhoff", "UC Santa Cruz",
     "Wissenschaftliche Ressource zur Traumforschung; beinhaltet Findings zu Barb Sanders, Engine Man u.a.",
     "Aktiv"],
    ["REF-006", "IASD – International Association for the Study of Dreams", "Organisation",
     "https://www.asdreams.org/", "",
     "Diverse Mitglieder", "IASD",
     "Internationale Vereinigung für Traumforschung; publiziert das Journal Dreaming.",
     "Aktiv"],
    ["REF-007", "APA Journal: Dreaming", "Fachzeitschrift",
     "https://www.apa.org/pubs/journals/drm", "",
     "Senior Editor: Kelly Bulkeley", "American Psychological Association",
     "Peer-reviewed Zeitschrift für Traumforschung; veröffentlicht SDDb-basierte Studien.",
     "Aktiv"],
    ["REF-008", "SDDb Guide", "Dokumentation",
     "https://bulkeley.org/guide-sleep-dream-database/", "",
     "Kelly Bulkeley", "SDDb",
     "Offizieller Leitfaden zur Nutzung der Sleep and Dream Database.",
     "Aktiv"],
    ["REF-009", "Pandemic COVID-19 Survey (DOI)", "Publikation",
     "https://doi.org/10.1037/drm0000146", "10.1037/drm0000146",
     "Michael Schredl, Kelly Bulkeley",
     "Central Institute of Mental Health / SDDb",
     "Dreaming and the COVID-19 Pandemic: A Survey in a U.S. Sample. Dreaming 30(3), 189-198.",
     "Publiziert 2020"],
    ["REF-010", "BLM Dreams Survey (DOI)", "Publikation",
     "https://doi.org/10.1007/s11089-021-00986-x", "10.1007/s11089-021-00986-x",
     "Kelly Bulkeley, Michael Schredl",
     "SDDb / Central Institute of Mental Health",
     "Dreams, Race, and the Black Lives Matter Movement. Pastoral Psychology 71, 29-41 (2022).",
     "Publiziert 2022"],
    ["REF-011", "Digital Dream Analysis (DOI)", "Publikation",
     "https://doi.org/10.1016/j.concog.2014.07.006", "10.1016/j.concog.2014.07.006",
     "Kelly Bulkeley", "SDDb",
     "Digital Dream Analysis: A Revised Method. Consciousness and Cognition 29, 17-30 (2014).",
     "Publiziert 2014"],
    ["REF-012", "9/11 Dreams Santa Clara (DOI)", "Publikation",
     "https://doi.org/10.1016/j.concog.2008.07.009", "10.1016/j.concog.2008.07.009",
     "Kelly Bulkeley, Tracey L. Kahan", "Santa Clara University",
     "The Impact of September 11 on Dreaming. Consciousness and Cognition 17(4), 1248-1256 (2008).",
     "Publiziert 2008"],
    ["REF-013", "Hall & Van de Castle (1966)", "Buch",
     "", "",
     "Calvin S. Hall, Robert L. Van de Castle", "Appleton-Century-Crofts",
     "The Content Analysis of Dreams. Klassisches Normwerk mit 1000 Traumberichten (981 im SDDb).",
     "Historisch 1966"],
    ["REF-014", "Dryad: DreamBank dataset", "Datensatz (Dryad)",
     "https://datadryad.org/dataset/doi:10.5061/dryad.r7sqv9scc",
     "10.5061/dryad.r7sqv9scc",
     "G. William Domhoff, Adam Schneider", "UC Santa Cruz",
     "DreamBank.net raw data auf Dryad veröffentlicht.",
     "Archiviert"],
]

# Zusätzlich alle Studien aus studies[] als Referenzen
for study in studies_data.get("studies", []):
    doi = study.get("doi", "") or ""
    doi_url = study.get("doi_url") or study.get("publication_url") or ""
    pub = study.get("publication", "") or ""
    researchers_list = study.get("researchers", [])
    refs_row = [
        f"REF-{len(refs)+1:03d}",
        study.get("name", ""),
        "Studie/Publikation",
        doi_url,
        doi,
        ", ".join(researchers_list),
        study.get("institution", ""),
        (pub + " — " if pub else "") + (study.get("description", "") or ""),
        f"{'Publiziert ' + str(study.get('year','')) if study.get('year') else 'Unbekannt'}",
    ]
    refs.append(refs_row)

for ref in refs:
    ws7.append(ref)

style_header_row(ws7)
freeze_header(ws7)
auto_width(ws7)

# ── Speichern ──────────────────────────────────────────────────────────────────
print(f"Speichere {OUTPUT} …")
wb.save(OUTPUT)
print(f"Fertig! Datei gespeichert: {OUTPUT}")
print(f"Größe: {OUTPUT.stat().st_size / 1024 / 1024:.1f} MB")
